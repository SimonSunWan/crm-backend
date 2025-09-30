from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.crud.dictionary import dictionary_enum_crud, dictionary_type_crud
from app.schemas.base import ApiResponse
from app.schemas.dictionary import (
    DictionaryEnumCreate,
    DictionaryEnumResponse,
    DictionaryEnumUpdate,
    DictionaryTypeCreate,
    DictionaryTypeResponse,
    DictionaryTypeUpdate,
    BatchImportDictionaryEnumRequest,
    BatchImportResult,
)

router = APIRouter()


# 字典类型相关接口
@router.get("/types", response_model=ApiResponse)
def get_dictionary_types(
    current: int = 1, size: int = 10, name: str = None, db: Session = Depends(get_db)
):
    """获取字典类型列表"""
    try:
        skip = (current - 1) * size

        # 构建查询
        query = db.query(dictionary_type_crud.model)

        # 添加搜索条件
        if name:
            query = query.filter(dictionary_type_crud.model.name.contains(name))

        # 获取总数
        total = query.count()

        # 获取分页数据
        types = query.offset(skip).limit(size).all()

        type_responses = [
            DictionaryTypeResponse.model_validate(type_obj) for type_obj in types
        ]

        response_data = {
            "records": type_responses,
            "total": total,
            "current": current,
            "size": size,
        }

        return ApiResponse(data=response_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取字典类型列表失败: {str(e)}")


@router.get("/by-code/{code}", response_model=ApiResponse)
def get_dictionary_by_code(code: str, db: Session = Depends(get_db)):
    """根据字典编码获取字典数据"""
    try:
        # 根据编码获取字典类型
        dictionary_type = dictionary_type_crud.get_by_code(db, code)
        if not dictionary_type:
            raise HTTPException(status_code=404, detail="字典类型不存在")

        # 构建级联树结构，只返回根级项
        tree = dictionary_enum_crud.build_cascade_tree(db, dictionary_type.id)

        enum_responses = [
            DictionaryEnumResponse.model_validate(enum_obj) for enum_obj in tree
        ]

        response_data = {
            "type": DictionaryTypeResponse.model_validate(dictionary_type),
            "enums": enum_responses,
        }

        return ApiResponse(data=response_data)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取字典数据失败: {str(e)}")


@router.get("/public/by-code/{code}", response_model=ApiResponse)
def get_dictionary_by_code_public(code: str, db: Session = Depends(get_db)):
    """根据字典编码获取字典数据（公开接口，无需认证）"""
    try:
        # 根据编码获取字典类型
        dictionary_type = dictionary_type_crud.get_by_code(db, code)
        if not dictionary_type:
            raise HTTPException(status_code=404, detail="字典类型不存在")

        # 构建级联树结构，只返回根级项
        tree = dictionary_enum_crud.build_cascade_tree(db, dictionary_type.id)

        enum_responses = [
            DictionaryEnumResponse.model_validate(enum_obj) for enum_obj in tree
        ]

        response_data = {
            "type": DictionaryTypeResponse.model_validate(dictionary_type),
            "enums": enum_responses,
        }

        return ApiResponse(data=response_data)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取字典数据失败: {str(e)}")


@router.post("/types", response_model=ApiResponse)
def create_dictionary_type(
    type_data: DictionaryTypeCreate, db: Session = Depends(get_db)
):
    """创建字典类型"""
    try:
        # 检查名称和编码是否已存在
        if dictionary_type_crud.get_by_name(db, type_data.name):
            raise HTTPException(status_code=400, detail="字典类型名称已存在")
        if dictionary_type_crud.get_by_code(db, type_data.code):
            raise HTTPException(status_code=400, detail="字典类型编码已存在")

        created_type = dictionary_type_crud.create(db, type_data.model_dump(by_alias=False))
        return ApiResponse(
            message="字典类型创建成功",
            data=DictionaryTypeResponse.model_validate(created_type),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"创建字典类型失败: {str(e)}")


@router.get("/types/{type_id}", response_model=ApiResponse)
def get_dictionary_type(type_id: int, db: Session = Depends(get_db)):
    """获取单个字典类型"""
    try:
        type_obj = dictionary_type_crud.get_or_404(db, type_id, "字典类型未找到")
        return ApiResponse(data=DictionaryTypeResponse.model_validate(type_obj))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取字典类型失败: {str(e)}")


@router.put("/types/{type_id}", response_model=ApiResponse)
def update_dictionary_type(
    type_id: int, type_update: DictionaryTypeUpdate, db: Session = Depends(get_db)
):
    """更新字典类型"""
    try:
        type_obj = dictionary_type_crud.get_or_404(db, type_id, "字典类型未找到")

        # 检查名称和编码是否重复
        if type_update.name and type_update.name != type_obj.name:
            if dictionary_type_crud.get_by_name(db, type_update.name):
                raise HTTPException(status_code=400, detail="字典类型名称已存在")
        if type_update.code and type_update.code != type_obj.code:
            if dictionary_type_crud.get_by_code(db, type_update.code):
                raise HTTPException(status_code=400, detail="字典类型编码已存在")

        updated_type = dictionary_type_crud.update(
            db, type_obj, type_update.model_dump(exclude_unset=True)
        )
        return ApiResponse(
            message="字典类型更新成功",
            data=DictionaryTypeResponse.model_validate(updated_type),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新字典类型失败: {str(e)}")


@router.delete("/types/{type_id}", response_model=ApiResponse)
def delete_dictionary_type(type_id: int, db: Session = Depends(get_db)):
    """删除字典类型"""
    try:
        type_obj = dictionary_type_crud.get_or_404(db, type_id, "字典类型未找到")
        dictionary_type_crud.delete(db, type_obj)
        return ApiResponse(message="字典类型删除成功")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除字典类型失败: {str(e)}")


# 字典枚举相关接口
@router.get("/enums", response_model=ApiResponse)
def get_dictionary_enums(type_id: int, db: Session = Depends(get_db)):
    """获取字典枚举列表"""
    try:
        tree = dictionary_enum_crud.build_cascade_tree(db, type_id)
        enum_responses = [
            DictionaryEnumResponse.model_validate(enum_obj) for enum_obj in tree
        ]

        return ApiResponse(
            data={
                "records": enum_responses,
                "total": len(enum_responses),
                "current": 1,
                "size": len(enum_responses),
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取字典枚举列表失败: {str(e)}")


@router.get("/enums/root", response_model=ApiResponse)
def get_root_dictionary_enums(type_id: int, db: Session = Depends(get_db)):
    """获取根级字典枚举列表"""
    try:
        root_enums = dictionary_enum_crud.get_root_enums(db, type_id)
        enum_responses = [
            DictionaryEnumResponse.model_validate(enum_obj) for enum_obj in root_enums
        ]

        return ApiResponse(data=enum_responses)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取根级字典枚举失败: {str(e)}")


@router.get("/enums/{parent_id}/children", response_model=ApiResponse)
def get_children_dictionary_enums(parent_id: int, db: Session = Depends(get_db)):
    """获取子级字典枚举列表"""
    try:
        children_enums = dictionary_enum_crud.get_children_by_parent_id(db, parent_id)
        enum_responses = [
            DictionaryEnumResponse.model_validate(enum_obj)
            for enum_obj in children_enums
        ]

        return ApiResponse(data=enum_responses)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取子级字典枚举失败: {str(e)}")


@router.post("/enums", response_model=ApiResponse)
def create_dictionary_enum(
    enum_data: DictionaryEnumCreate, db: Session = Depends(get_db)
):
    """创建字典枚举"""
    try:
        # 检查键值是否已存在
        if dictionary_enum_crud.get_by_type_id_and_key(
            db, enum_data.type_id, enum_data.key_value
        ):
            raise HTTPException(status_code=400, detail="键值已存在")

        # 计算层级和路径
        level = 1
        path = enum_data.key_value

        if enum_data.parent_id:
            parent_enum = dictionary_enum_crud.get(db, enum_data.parent_id)
            if not parent_enum:
                raise HTTPException(status_code=400, detail="父级枚举不存在")
            level = (parent_enum.level or 0) + 1
            path = (
                f"{parent_enum.path}/{enum_data.key_value}"
                if parent_enum.path
                else f"{parent_enum.key_value}/{enum_data.key_value}"
            )

        # 创建枚举数据
        enum_dict = enum_data.model_dump(by_alias=False)
        enum_dict["level"] = level
        enum_dict["path"] = path

        created_enum = dictionary_enum_crud.create(db, enum_dict)
        return ApiResponse(
            message="字典枚举创建成功",
            data=DictionaryEnumResponse.model_validate(created_enum),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"创建字典枚举失败: {str(e)}")


@router.get("/enums/{enum_id}", response_model=ApiResponse)
def get_dictionary_enum(enum_id: int, db: Session = Depends(get_db)):
    """获取单个字典枚举"""
    try:
        enum_obj = dictionary_enum_crud.get_or_404(db, enum_id, "字典枚举未找到")
        return ApiResponse(data=DictionaryEnumResponse.model_validate(enum_obj))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取字典枚举失败: {str(e)}")


@router.put("/enums/{enum_id}", response_model=ApiResponse)
def update_dictionary_enum(
    enum_id: int, enum_update: DictionaryEnumUpdate, db: Session = Depends(get_db)
):
    """更新字典枚举"""
    try:
        enum_obj = dictionary_enum_crud.get_or_404(db, enum_id, "字典枚举未找到")

        # 检查键值是否重复
        if enum_update.key_value and enum_update.key_value != enum_obj.key_value:
            if dictionary_enum_crud.get_by_type_id_and_key(
                db, enum_obj.type_id, enum_update.key_value
            ):
                raise HTTPException(status_code=400, detail="键值已存在")

        updated_enum = dictionary_enum_crud.update(
            db, enum_obj, enum_update.model_dump(exclude_unset=True, by_alias=False)
        )
        return ApiResponse(
            message="字典枚举更新成功",
            data=DictionaryEnumResponse.model_validate(updated_enum),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新字典枚举失败: {str(e)}")


@router.delete("/enums/{enum_id}", response_model=ApiResponse)
def delete_dictionary_enum(enum_id: int, db: Session = Depends(get_db)):
    """删除字典枚举（级联删除子级）"""
    try:
        dictionary_enum_crud.delete_cascade(db, enum_id)
        return ApiResponse(message="字典枚举删除成功")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除字典枚举失败: {str(e)}")


@router.post("/enums/batch-import", response_model=ApiResponse)
def batch_import_dictionary_enums(
    request_data: BatchImportDictionaryEnumRequest, db: Session = Depends(get_db)
):
    """批量导入字典枚举"""
    try:
        result = dictionary_enum_crud.batch_import(db, request_data.type_id, request_data.data)
        return ApiResponse(data=result)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"批量导入失败: {str(e)}")


@router.get("/enums/template/{type_id}")
def download_dictionary_enum_template(type_id: int, db: Session = Depends(get_db)):
    """下载字典枚举导入模板"""
    try:
        from fastapi.responses import StreamingResponse
        import io
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # 检查字典类型是否存在
        dictionary_type = dictionary_type_crud.get(db, type_id)
        if not dictionary_type:
            raise HTTPException(status_code=404, detail="字典类型不存在")
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        # 清理工作表名称，移除不允许的字符
        safe_name = dictionary_type.name.replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_').replace('[', '_').replace(']', '_')
        ws.title = f"{safe_name}导入模板"
        
        # 设置标题行
        headers = ["枚举编码", "枚举名称", "排序", "父级编码", "层级"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置所有列都一样宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # 添加简单示例数据
        example_data = [
            ["cy", "乘用车", 1, "", 1],
            ["cy_sedan", "轿车", 1, "cy", 2]
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 返回文件流
        filename = f"{dictionary_type.name}_导入模板.xlsx"
        # 对文件名进行URL编码以支持中文
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"生成模板失败: {str(e)}")
